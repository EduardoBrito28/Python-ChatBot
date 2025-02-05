# text.py
from PyPDF2 import PdfReader
from langchain.text_splitter import CharacterTextSplitter
from utils import db
import openpyxl
import os


def process_files(files):
    text = ""
    for file in files:
        if isinstance(file, str):
            # Se for um caminho de arquivo no sistema
            file_extension = os.path.splitext(file)[1].lower()
            with open(file, "rb") as f:
                if file_extension == ".pdf":
                    pdf = PdfReader(f)
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text
                elif file_extension == ".xlsx":
                    wb = openpyxl.load_workbook(f)
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for row in ws.iter_rows(values_only=True):
                            text += " ".join([str(cell) for cell in row if cell]) + "\n"
        else:
            # Se for um objeto de upload do Streamlit
            if file.name.endswith(".pdf"):
                pdf = PdfReader(file)
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text
            elif file.name.endswith(".xlsx"):
                wb = openpyxl.load_workbook(file)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text


def create_text_chunks(text):
    text_splitter = CharacterTextSplitter(
        separator="\n", chunk_size=1500, chunk_overlap=300, length_function=len
    )
    chunks = text_splitter.split_text(text)
    return chunks


def save_document_metadata(filename, content, file_hash):
    document_data = {
        "filename": filename,
        "content": content,
        "hash": file_hash,
    }
    db.documents_collection.insert_one(document_data)


def document_exists(file_hash):
    return db.documents_collection.find_one({"hash": file_hash}) is not None
